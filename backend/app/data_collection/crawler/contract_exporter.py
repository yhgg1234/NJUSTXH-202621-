# -*- coding: utf-8 -*-
"""
contract_exporter.py —— 把爬虫抽取出的 JD 记录导出为符合《2.1/2.2 上游数据交付契约 v1.0.0》的 xlsx
=========================================================================================================
与现有 JD数据_契约版.xlsx 格式完全一致(37 固定列 + 月度分表 + 采集日志 + 质量报告摘要)。

输入 records: list[dict], 每条含:
    platform        来源平台(中文, 如 "BOSS直聘" / "国聘")
    url             原始可核验链接
    job_title       岗位名称
    company         公司名称
    industry        所属行业
    city            工作城市
    responsibilities 岗位职责原文
    requirements     任职要求原文
    raw_skills      原始技能词
    tech_stack      技术栈
    certificates    证书要求
    education       学历要求
    experience      经验要求
    project_exp     项目经验要求
    external_id     平台原始 ID(可选, 取不到留空)
    crawled_month   采集月份 "YYYY-MM"(用于 crawled_at 与月度分表归档)

本模块只负责"格式化 + 溯源 + 写文件", 不做网络请求。
合规护栏: published_at 一律置 null(不猜测), extracted_* 置 [] / null 留给 2.1, 个人信息过滤在 filters.py 完成。
"""
import re, json, hashlib, unicodedata
from datetime import datetime, timezone, timedelta
import pandas as pd
from openpyxl import Workbook

TZ = timezone(timedelta(hours=8))
NOW = datetime.now(TZ).strftime("%Y-%m-%dT%H:%M:%S+08:00")

PLATFORM_SLUG = {
    "智联招聘": "zhaopin", "前程无忧": "51job", "猎聘": "liepin",
    "BOSS直聘": "boss", "BOSS直聘(店长)": "boss-dianzhang",
    "BOSS直聘官方(字节校招)": "boss-bytedance", "58同城": "58tongcheng",
    "CDA数据分析师官网": "cda", "中国公共招聘网": "mohrss", "中金招聘官网": "cicc",
    "企业官网": "company-site", "企查查招聘": "qcc", "全知招聘": "quanzhi",
    "千万岗位库(聚合)": "qianwanjob", "国聘": "iguopin",
    "宁波人才网/公共招聘网": "nbrc", "广州人才网": "gzrc", "应届生求职网": "yingjiesheng",
    "成都人才网": "rc114", "牛客网(字节校招)": "nowcoder-bytedance", "禾蛙": "hewa",
    "重庆人社": "cqhrss", "高校人才网": "gaoxiaojob", "高校就业网": "ncss",
    "高校就业网(河南工业大学)": "haut", "齐鲁人才网": "qlrc",
}
EXT_PATTERNS = [
    r"/jobdetail/([A-Za-z0-9]{8,})",
    r"jobs\.51job\.com/[^/]+/(\d{6,})\.html",
    r"/job/([0-9a-f]{12,})",
    r"jobAdId=([0-9a-f-]{8,})",
    r"/id-(\d{4,})",
    r"(?:jobId|positionId|job_id)=([A-Za-z0-9-]{6,})",
]
ALL_COLS = ["schema_version", "document_type", "jd_id", "source_platform",
    "external_id", "url", "published_at", "crawled_at", "first_seen_at",
    "last_seen_at", "content_hash", "language", "job_title_raw", "company_raw",
    "industry_raw", "city_raw", "job_level_raw", "responsibilities",
    "requirements", "raw_text", "extracted_entities_json",
    "extracted_relations_json", "extracted_events_json", "extraction_model",
    "model_version", "prompt_version", "extracted_at", "overall_confidence",
    "needs_human_review", "quality_issues_json", "legacy_jd_id",
    "education_required_raw", "experience_required_raw",
    "project_experience_raw", "raw_skills", "tech_stack", "certificates_raw"]

CJK = re.compile(r"[一-鿿]")

def sha256_hex(s):
    return hashlib.sha256(s.encode("utf-8")).hexdigest()

def norm_text_for_hash(s):
    s = unicodedata.normalize("NFKC", s)
    return re.sub(r"\s+", " ", s).strip().lower()

def norm_url(u):
    u = (u or "").strip()
    m = re.match(r"(https?://)([^/]+)(.*)", u, re.I)
    if not m:
        return u
    path = m.group(3)
    if path == "/":
        path = ""
    return m.group(1).lower() + m.group(2).lower() + path

def extract_external_id(url):
    for pat in EXT_PATTERNS:
        m = re.search(pat, url or "")
        if m:
            return m.group(1)
    return ""

def detect_lang(*texts):
    t = "".join(texts)
    if not t:
        return "zh-CN"
    cjk = len(CJK.findall(t))
    return "zh-CN" if cjk >= max(1, len(t) * 0.2) else "en-US"

def detect_level(title):
    for kw, lv in [("实习", "实习"), ("初级", "初级"), ("助理工程师", "初级"),
                   ("中级", "中级"), ("高级", "高级"), ("资深", "高级"), ("专家", "专家")]:
        if kw in title:
            return lv
    return ""

def slug_of(plat):
    return PLATFORM_SLUG.get(plat, re.sub(r"[^a-z0-9]+", "-", plat.lower()).strip("-") or "unknown")

def clean(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none", "暂无", "未知", "--", "nan") else s


def build_record(raw):
    """raw: dict(见模块 docstring)。返回 37 列契约 dict。"""
    plat = clean(raw.get("platform"))
    url = clean(raw.get("url"))
    title = clean(raw.get("job_title"))
    company = clean(raw.get("company"))
    resp = clean(raw.get("responsibilities"))
    req = clean(raw.get("requirements"))
    crawled_month = clean(raw.get("crawled_month")) or datetime.now(TZ).strftime("%Y-%m")
    ext = clean(raw.get("external_id")) or extract_external_id(url)
    raw_text = (resp + "\n\n" + req).strip()
    content_hash = "sha256:" + sha256_hex(norm_text_for_hash(raw_text + "|" + title + "|" + company))

    slug = slug_of(plat)
    # jd_id: 外部ID(唯一) -> 唯一URL的SHA-256 -> 正文SHA-256
    if ext:
        jd_id = f"{slug}:{ext}"
    elif url and norm_url(url):
        jd_id = f"{slug}:sha256:{sha256_hex(norm_url(url))[:32]}"
    else:
        jd_id = f"{slug}:sha256:{sha256_hex(norm_text_for_hash(plat + '|' + raw_text + '|' + title + '|' + company))[:32]}"

    crawled = f"{crawled_month}-01T00:00:00+08:00"
    issues = []
    if not ext:
        issues.append("missing_external_id")
    return {
        "schema_version": "1.0.0",
        "document_type": "job_description",
        "jd_id": jd_id,
        "source_platform": plat,
        "external_id": ext,
        "url": url,
        "published_at": "",                # 未知 -> null, 契约禁止猜测
        "crawled_at": crawled,
        "first_seen_at": crawled,
        "last_seen_at": crawled,
        "content_hash": content_hash,
        "language": detect_lang(title, resp, req),
        "job_title_raw": title,
        "company_raw": company,
        "industry_raw": clean(raw.get("industry")),
        "city_raw": clean(raw.get("city")),
        "job_level_raw": detect_level(title),
        "responsibilities": resp,
        "requirements": req,
        "raw_text": raw_text,
        "extracted_entities_json": "[]",    # 待 2.1 填充
        "extracted_relations_json": "[]",
        "extracted_events_json": "[]",
        "extraction_model": "",
        "model_version": "",
        "prompt_version": "",
        "extracted_at": "",
        "overall_confidence": "",
        "needs_human_review": "",
        "quality_issues_json": json.dumps(issues, ensure_ascii=False),
        "legacy_jd_id": "",
        "education_required_raw": clean(raw.get("education")),
        "experience_required_raw": clean(raw.get("experience")),
        "project_experience_raw": clean(raw.get("project_exp")),
        "raw_skills": clean(raw.get("raw_skills")),
        "tech_stack": clean(raw.get("tech_stack")),
        "certificates_raw": clean(raw.get("certificates")),
    }


def export(records, dst_xlsx, batch_id=None):
    """records: list[dict]; 写出契约版 xlsx, 返回 (out_df, report)。"""
    out_rows = []
    seen = set()
    for raw in records:
        rec = build_record(raw)
        # jd_id 兜底唯一
        if rec["jd_id"] in seen:
            rec["jd_id"] = f"{rec['jd_id']}-{len(seen)}"
        seen.add(rec["jd_id"])
        out_rows.append(rec)
    out = pd.DataFrame(out_rows, columns=ALL_COLS)

    # 月度分表
    out["__m"] = out["crawled_at"].astype(str).str[:7]
    months = sorted(out["__m"].dropna().unique().tolist())

    # 采集日志(平台 × 月份)
    logs = []
    i = 0
    for (ym, plat), g in out.groupby(["__m", "source_platform"], sort=True):
        i += 1
        slug = slug_of(plat)
        logs.append({
            "log_id": f"log-{i:04d}",
            "batch_id": batch_id or f"collect-{slug}-{str(ym).replace('-', '')}-01",
            "source_platform": plat,
            "source_slug": slug,
            "crawl_partition": ym,
            "crawled_at": f"{ym}-01T00:00:00+08:00",
            "records": len(g),
            "duplicates_removed_in_batch": "",
            "jd_id_rule": "external_id > url_sha256 > content_sha256",
            "note": "采集时间精确到月(按 crawled_month 归档),日用01占位; published_at 置 null",
            "logged_at": NOW,
        })
    log_df = pd.DataFrame(logs)

    report = {
        "schema_version": "1.0.0",
        "batch_id": batch_id or "collect-crawler-demo",
        "total_records": int(len(out)),
        "valid_records": int(len(out)),
        "duplicate_records": 0,
        "missing_published_at": int(len(out)),
        "missing_job_title": int((out["job_title_raw"].fillna("") == "").sum()),
        "source_platforms": int(out["source_platform"].nunique()),
        "unique_job_titles": int(out["job_title_raw"].nunique()),
        "jd_id_unique": bool(out["jd_id"].is_unique),
        "generated_at": NOW,
    }

    # ---- 写 Excel ----
    wb = Workbook(write_only=True)

    def write_sheet(name, header, rows, widths=None):
        ws = wb.create_sheet(name)
        if widths:
            from openpyxl.utils import get_column_letter
            for idx, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(idx)].width = w
        ws.append(header)
        for row in rows:
            ws.append(row)

    widths = [10, 14, 40, 18, 22, 40, 20, 20, 20, 20, 24, 9, 24, 22, 18, 10,
              12, 50, 50, 60, 12, 12, 12, 12, 12, 12, 18, 10, 12, 26,
              12, 14, 12, 24, 30, 30, 16]
    write_sheet("汇总(全部)", ALL_COLS, out[ALL_COLS].itertuples(index=False, name=None), widths)
    write_sheet("采集日志", list(log_df.columns), log_df.itertuples(index=False, name=None),
                [10, 30, 22, 18, 12, 22, 9, 22, 36, 52, 22])
    qa_rows = [(k, json.dumps(v, ensure_ascii=False) if isinstance(v, (dict, list)) else v)
               for k, v in report.items()]
    write_sheet("质量报告摘要", ["指标", "值"], qa_rows, [28, 90])
    for ms in months:
        sub = out[out["__m"] == ms]
        write_sheet(ms, ALL_COLS, sub[ALL_COLS].itertuples(index=False, name=None))
    out.drop(columns=["__m"], inplace=True)
    wb.save(dst_xlsx)

    return out, report
